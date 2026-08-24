import io
import json
import re
import pandas as pd
from datetime import datetime, timedelta
from bot_system.db.connection import get_db_connection

def get_dashboard_metrics():
    """Requirement 5: Dashboard metrics (Today's count, payout due totals for 6-day & 1-day, unique UPI count). Any status except 'paid' and 'rejected' is considered due."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # 1. Today's incoming codes count & list
            cursor.execute("""
                SELECT id, phone_number, gift_card_name, gift_card_code, payment_details, total_amount, status, created_at
                FROM submissions 
                WHERE DATE(created_at) = CURRENT_DATE()
                ORDER BY created_at DESC
            """)
            today_list = list(cursor.fetchall())
            today_codes = 0
            for t in today_list:
                t["created_at_str"] = t["created_at"].strftime("%H:%M")
                codes = [c.strip() for c in re.split(r'[\n,\s;]+', str(t.get("gift_card_code") or "")) if c.strip()]
                today_codes += len(codes)

            # 2. Payout Due Totals (status != 'paid' AND status != 'rejected')
            cursor.execute("""
                SELECT 
                    SUM(CASE WHEN payout_term_days = 6 THEN total_amount ELSE 0 END) as term_6_total,
                    SUM(CASE WHEN payout_term_days = 1 THEN total_amount ELSE 0 END) as term_1_total,
                    SUM(total_amount) as total_payout,
                    COUNT(DISTINCT payment_details) as unique_upis,
                    COUNT(DISTINCT phone_number) as unique_suppliers
                FROM submissions 
                WHERE status NOT IN ('paid', 'rejected')
            """)
            payouts = cursor.fetchone()

            # 3. Code Status Summary (counts actual codes)
            cursor.execute("""
                SELECT gift_card_code, status, c_status, code_statuses
                FROM submissions
            """)
            all_subs = cursor.fetchall()
            total_unpaid = 0
            total_listed = 0
            total_sold = 0

            for sub in all_subs:
                sub_st = str(sub.get("status") or "unpaid").lower()
                row_cst = str(sub.get("c_status") or "unsold").lower()
                if row_cst == "ununsold": row_cst = "unsold"

                # Parse code_statuses JSON if present
                st_map = {}
                if sub.get("code_statuses"):
                    raw_st = sub["code_statuses"]
                    if isinstance(raw_st, str):
                        try: st_map = json.loads(raw_st)
                        except: st_map = {}
                    elif isinstance(raw_st, dict):
                        st_map = raw_st

                codes = [c.strip() for c in re.split(r'[\n,\s;]+', str(sub.get("gift_card_code") or "")) if c.strip()]
                if not codes:
                    continue

                for c in codes:
                    meta = st_map.get(c)
                    code_st = row_cst
                    if isinstance(meta, dict):
                        code_st = str(meta.get("c_status") or row_cst).lower()
                    elif isinstance(meta, str):
                        code_st = str(meta or row_cst).lower()
                    if code_st == "ununsold": code_st = "unsold"

                    # Count unpaid (submission status not paid and not rejected)
                    if sub_st not in ('paid', 'rejected'):
                        total_unpaid += 1

                    # Count listed & sold based on code c_status or submission status
                    if code_st == 'listed' or sub_st == 'listed':
                        total_listed += 1
                    if code_st == 'sold' or sub_st == 'sold':
                        total_sold += 1

            return {
                "today_codes": today_codes,
                "today_list": today_list,
                "term_6_total": float(payouts["term_6_total"] or 0),
                "term_1_total": float(payouts["term_1_total"] or 0),
                "total_payout": float(payouts["total_payout"] or 0),
                "unique_upis": payouts["unique_upis"] or 0,
                "unique_suppliers": payouts["unique_suppliers"] or 0,
                "total_unpaid": total_unpaid,
                "total_listed": total_listed,
                "total_sold": total_sold,
            }
    finally:
        conn.close()

from bot_system.services.old_csv_checker import is_code_in_old_csv

def get_codes_inventory():
    """Requirement 3.0, 3.1 & 6: Codes inventory combining submissions and blacklisted scammers with RED highlighting and DB stored Old CSV flag."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            # 1. Fetch Submissions
            cursor.execute("""
                SELECT id, phone_number, gift_card_name, gift_card_code, payment_details,
                       total_amount, COALESCE(currency, 'Rs.') as currency, status, COALESCE(c_status, 'unsold') as c_status, code_statuses, payout_term_days, card_type, platform, slot, order_id, created_at,
                       TIMESTAMPDIFF(HOUR, created_at, NOW()) as age_hours, 0 as is_scammer, '' as scam_reason, COALESCE(in_old_db, 0) as in_old_db
                FROM submissions
            """)
            submissions = list(cursor.fetchall())
            for r in submissions:
                r["row_source"] = "submission"

            # 2. Fetch Scammers & Duplicate Blacklist Records
            cursor.execute("""
                SELECT id, phone_number, COALESCE(gift_card_name, 'FLAGGED / DUPLICATE') as gift_card_name, flagged_code as gift_card_code,
                       payment_details, COALESCE(total_amount, 0.00) as total_amount, 'Rs.' as currency, COALESCE(status, 'flagged') as status, COALESCE(c_status, 'unsold') as c_status, NULL as code_statuses, created_at,
                       0 as age_hours, 1 as is_scammer, reason as scam_reason,
                       COALESCE(platform, '') as platform, COALESCE(slot, '') as slot, COALESCE(order_id, '') as order_id, COALESCE(in_old_db, 0) as in_old_db
                FROM scammers
            """)
            scammers = list(cursor.fetchall())
            for r in scammers:
                r["row_source"] = "scammer"

            all_records = submissions + scammers
            all_records.sort(key=lambda x: x["created_at"], reverse=True)

            for r in all_records:
                if hasattr(r["created_at"], "strftime"):
                    r["created_at_str"] = r["created_at"].strftime("%Y-%m-%d %H:%M")
                    r["created_at_iso"] = r["created_at"].isoformat()
                else:
                    r["created_at_str"] = str(r.get("created_at") or "")
                    r["created_at_iso"] = str(r.get("created_at") or "")

                # Parse per-code status mapping
                statuses_map = {}
                if r.get("code_statuses"):
                    raw_st = r["code_statuses"]
                    if isinstance(raw_st, str):
                        try: statuses_map = json.loads(raw_st)
                        except: statuses_map = {}
                    elif isinstance(raw_st, dict):
                        statuses_map = raw_st

                raw_cst = r.get("c_status") or "unsold"
                default_row_c_status = "unsold" if raw_cst == "ununsold" else raw_cst
                default_row_platform = r.get("platform") or ""
                default_row_slot = r.get("slot") or ""
                codes = [c.strip() for c in re.split(r'[\n,\s;]+', str(r.get("gift_card_code") or "")) if c.strip()]
                parsed_code_details = []
                for c in codes:
                    meta = statuses_map.get(c)
                    if isinstance(meta, dict):
                        code_st = meta.get("c_status") or default_row_c_status
                        code_plat = meta.get("platform") or default_row_platform
                        code_slot = meta.get("slot") or default_row_slot
                    elif isinstance(meta, str):
                        code_st = meta or default_row_c_status
                        code_plat = default_row_platform
                        code_slot = default_row_slot
                    else:
                        code_st = default_row_c_status
                        code_plat = default_row_platform
                        code_slot = default_row_slot

                    if code_st == "ununsold":
                        code_st = "unsold"

                    parsed_code_details.append({
                        "code": c,
                        "c_status": code_st,
                        "platform": code_plat,
                        "slot": code_slot
                    })
                r["code_details"] = parsed_code_details

                # If in_old_db not already flagged in DB, fallback to quick CSV set lookup
                if not r.get("in_old_db"):
                    if any(is_code_in_old_csv(c) for c in codes):
                        r["in_old_db"] = 1
                else:
                    r["in_old_db"] = int(r["in_old_db"])

                # Check if card_type explicitly set to OLD in message, otherwise check age
                card_t = str(r.get("card_type") or "").upper()
                if card_t == "OLD":
                    r["is_new"] = 0
                elif card_t == "NEW":
                    r["is_new"] = 1
                else:
                    r["is_new"] = 1 if r["age_hours"] < 24 else 0
                if r["is_scammer"]:
                    r["aging_alert"] = "SCAMMER_RED"
                elif r["age_hours"] >= 24 and r["status"] != "sold":
                    r["aging_alert"] = "RED" if r["age_hours"] >= 48 else "YELLOW"
                else:
                    r["aging_alert"] = "NORMAL"
            return all_records
    finally:
        conn.close()

import json

def get_app_config():
    """Fetch current dynamic config for platforms and slots from DB."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT config_key, config_value FROM app_config")
            rows = cursor.fetchall()
            config = {"platforms": ["test1", "test2", "test3"], "slots": ["test1", "test2", "test3"]}
            for row in rows:
                key = row["config_key"]
                val = row["config_value"]
                if isinstance(val, str):
                    try: val = json.loads(val)
                    except: pass
                if key in config and isinstance(val, list):
                    config[key] = val
            return config
    finally:
        conn.close()

def update_app_config(key: str, items: list):
    """Save updated platform or slot options list to DB."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            json_val = json.dumps(items)
            cursor.execute("""
                INSERT INTO app_config (config_key, config_value)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE config_value = %s
            """, (key, json_val, json_val))
            return True
    finally:
        conn.close()

def get_scammers_list():
    """Requirement 3.1: Duplicate cards & rejected scammer accounts blacklist."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM scammers ORDER BY created_at DESC")
            records = cursor.fetchall()
            for r in records:
                r["created_at_str"] = r["created_at"].strftime("%Y-%m-%d %H:%M")
            return records
    finally:
        conn.close()

def export_payout_excel(days_offset: int = 6, start_date: str = None, end_date: str = None, mode: str = "summary"):
    """
    Advanced Payout Excel Exporter:
    - Day-wise daily sheets or single summary sheet
    - Grouped by Phone Number
    - Highlights repeated Phone numbers and repeated UPI IDs with custom fills
    - Individual code breakdown with Sold/Unsold/Listed/Paid/Rejected status tags
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    conn = get_db_connection()
    try:
        query_conditions = ["status NOT IN ('paid', 'rejected')"]
        params = []

        if start_date and end_date:
            query_conditions.append("DATE(created_at) BETWEEN %s AND %s")
            params.extend([start_date, end_date])
        elif start_date:
            query_conditions.append("DATE(created_at) >= %s")
            params.append(start_date)
        elif days_offset is not None:
            target_date = (datetime.now() - timedelta(days=int(days_offset))).strftime("%Y-%m-%d")
            query_conditions.append("DATE(created_at) <= %s")
            params.append(target_date)

        where_clause = " AND ".join(query_conditions)
        sql = f"""
            SELECT 
                id, phone_number, gift_card_name, gift_card_code,
                payment_method, payment_details, total_amount, currency,
                card_type, platform, slot, order_id, status, code_statuses,
                payout_term_days, created_at
            FROM submissions
            WHERE {where_clause}
            ORDER BY DATE(created_at) DESC, phone_number ASC, created_at ASC
        """

        with conn.cursor() as cursor:
            cursor.execute(sql, tuple(params))
            records = cursor.fetchall()

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default blank sheet

        # Frequency counters for highlighting duplicate phone / UPI across dataset
        phone_counts = {}
        upi_counts = {}
        for r in records:
            p = str(r.get("phone_number") or "").strip()
            u = str(r.get("payment_details") or "").strip()
            if p: phone_counts[p] = phone_counts.get(p, 0) + 1
            if u: upi_counts[u] = upi_counts.get(u, 0) + 1

        # Color Styles
        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid") # Excel Green
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="107C41")
        subtotal_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)

        # Highlight fills
        dup_phone_fill = PatternFill(start_color="FFD8D8", end_color="FFD8D8", fill_type="solid") # Soft Red
        dup_upi_fill = PatternFill(start_color="FFE8CC", end_color="FFE8CC", fill_type="solid")   # Soft Orange/Amber

        # Code Status Fills
        status_fills = {
            "sold": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),      # Green
            "listed": PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),    # Blue
            "unsold": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),    # Yellow
            "unpaid": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
            "paid": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            "rejected": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        }

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        total_top_bottom_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000')
        )

        headers = [
            "Date", "Phone Number", "Supplier UPI / Payment", "Order ID",
            "Card Name", "Code", "Code Status", "Platform", "Slot", 
            "Term", "Amount (₹)", "Submission Status"
        ]

        def format_sheet(ws, sheet_records, sheet_title):
            ws.views.sheetView[0].showGridLines = True
            ws.append([sheet_title])
            ws.cell(1, 1).font = title_font
            ws.append([]) # Blank row

            ws.append(headers)
            header_row_idx = 3
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center" if col_idx not in [2, 3, 6] else "left", vertical="center")
                cell.border = thin_border

            current_row = 4
            grand_total_amount = 0

            for r in sheet_records:
                phone = str(r.get("phone_number") or "").strip()
                upi = str(r.get("payment_details") or "").strip()
                created_d = r["created_at"].strftime("%Y-%m-%d %H:%M") if r.get("created_at") else ""
                order_id = r.get("order_id") or "—"
                card_name = r.get("gift_card_name") or "—"
                platform = r.get("platform") or "—"
                slot = r.get("slot") or "—"
                term = f"{r.get('payout_term_days', 6)}D"
                amount = float(r.get("total_amount") or 0.0)
                sub_status = str(r.get("status") or "unpaid").upper()
                grand_total_amount += amount

                # Parse individual codes and their c_status
                code_statuses = {}
                if r.get("code_statuses"):
                    try:
                        raw_cs = r["code_statuses"]
                        code_statuses = json.loads(raw_cs) if isinstance(raw_cs, str) else raw_cs
                    except:
                        pass

                raw_codes = [c.strip() for c in re.split(r'[\n,\s;]+', str(r.get("gift_card_code") or "")) if c.strip()]
                if not raw_codes:
                    raw_codes = ["—"]

                # Write each code line for complete transparency
                for idx, code_item in enumerate(raw_codes):
                    c_status = str(code_statuses.get(code_item, {}).get("c_status") if isinstance(code_statuses.get(code_item), dict) else code_statuses.get(code_item) or "unsold").lower()
                    if c_status in ['ununsold', '']: c_status = 'unsold'

                    row_data = [
                        created_d if idx == 0 else "",
                        phone if idx == 0 else "",
                        upi if idx == 0 else "",
                        order_id if idx == 0 else "",
                        card_name,
                        code_item,
                        c_status.upper(),
                        platform,
                        slot,
                        term if idx == 0 else "",
                        amount if idx == 0 else 0.0,
                        sub_status if idx == 0 else ""
                    ]
                    ws.append(row_data)

                    # Cell Styling & Duplicate Highlighting
                    for c_i in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=c_i)
                        cell.font = regular_font
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")

                        # Align numbers
                        if c_i == 11:
                            cell.number_format = '₹#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")

                        # Highlight Duplicate Phone
                        if c_i == 2 and phone and phone_counts.get(phone, 0) > 1 and idx == 0:
                            cell.fill = dup_phone_fill
                            cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")

                        # Highlight Duplicate UPI
                        if c_i == 3 and upi and upi_counts.get(upi, 0) > 1 and idx == 0:
                            cell.fill = dup_upi_fill
                            cell.font = Font(name="Calibri", size=10, bold=True, color="9C6500")

                        # Highlight Code Status Pill
                        if c_i == 7:
                            c_fill = status_fills.get(c_status, None)
                            if c_fill:
                                cell.fill = c_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            cell.font = Font(name="Calibri", size=10, bold=True)

                    current_row += 1

            # Summary Footer Row
            ws.append([])
            current_row += 1
            ws.append(["", "", "", "", "", "GRAND TOTAL DUE", "", "", "", "", grand_total_amount, ""])
            summary_row = current_row
            for c_i in range(1, len(headers) + 1):
                cell = ws.cell(row=summary_row, column=c_i)
                cell.font = bold_font
                cell.fill = subtotal_fill
                cell.border = total_top_bottom_border
                if c_i == 11:
                    cell.number_format = '₹#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    if cell.row > 1 and len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Build Sheets
        if not records:
            ws = wb.create_sheet(title="Payout Summary")
            format_sheet(ws, [], "Payout Report (No Records Found)")
        else:
            # 1. Main Master Summary Sheet
            ws_master = wb.create_sheet(title="All Due Payouts")
            format_sheet(ws_master, records, "Master Payout Export - Grouped by Supplier")

            # 2. Daily Tabs (Grouped Day-wise)
            records_by_day = {}
            for r in records:
                d_str = r["created_at"].strftime("%Y-%m-%d") if r.get("created_at") else "No_Date"
                records_by_day.setdefault(d_str, []).append(r)

            for d_str, day_recs in records_by_day.items():
                tab_title = d_str[:31]  # Excel tab name max 31 chars
                ws_day = wb.create_sheet(title=tab_title)
                format_sheet(ws_day, day_recs, f"Payout Due for Date: {d_str}")

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    finally:
        conn.close()

def get_submission_full_details(submission_id: int):
    """Fetch complete raw and parsed audit details of a submission."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM submissions WHERE id = %s", (submission_id,))
            res = cursor.fetchone()
            if res:
                res["created_at_str"] = res["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                res["updated_at_str"] = res["updated_at"].strftime("%Y-%m-%d %H:%M:%S")
                res["row_source"] = "submission"
            return res
    finally:
        conn.close()

def get_scammer_full_details(scammer_id: int):
    """Fetch full details of a scammer record."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT * FROM scammers WHERE id = %s", (scammer_id,))
            res = cursor.fetchone()
            if res:
                res["created_at_str"] = res["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                res["gift_card_name"] = "FLAGGED / DUPLICATE"
                res["gift_card_code"] = res.get("flagged_code") or "—"
                res["total_amount"] = 0
                res["payout_term_days"] = None
                res["raw_message"] = res.get("reason") or "—"
                res["status"] = res.get("status") or "flagged"
                res["row_source"] = "scammer"
            return res
    finally:
        conn.close()


def get_all_duplicate_conflicts():
    """Detects duplicate gift card codes across all non-rejected submissions and scammers."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT id, phone_number, gift_card_name, gift_card_code, total_amount, currency, status, created_at, 'submission' as row_source
                FROM submissions
                WHERE status != 'rejected'
            """)
            submissions = list(cursor.fetchall())

            cursor.execute("""
                SELECT id, phone_number, COALESCE(gift_card_name, 'FLAGGED / DUPLICATE') as gift_card_name, flagged_code as gift_card_code,
                       total_amount, 'Rs.' as currency, status, created_at, 'scammer' as row_source
                FROM scammers
                WHERE status != 'rejected'
            """)
            scammers = list(cursor.fetchall())

        code_map = {}
        for r in submissions + scammers:
            created_str = r["created_at"].strftime("%Y-%m-%d %H:%M") if hasattr(r["created_at"], "strftime") else str(r.get("created_at") or "")
            raw_gc = str(r.get("gift_card_code") or "")
            codes = [c.strip() for c in re.split(r'[\n,\s;]+', raw_gc) if c.strip()]
            for c in codes:
                if not c:
                    continue
                if c not in code_map:
                    code_map[c] = []
                code_map[c].append({
                    "id": r["id"],
                    "phone_number": r.get("phone_number") or "",
                    "gift_card_name": r.get("gift_card_name") or "",
                    "all_codes": codes,
                    "code_count": len(codes),
                    "total_amount": float(r.get("total_amount") or 0),
                    "currency": r.get("currency") or "Rs.",
                    "status": r.get("status") or "",
                    "created_at_str": created_str,
                    "row_source": r.get("row_source")
                })

        duplicates = []
        for code, entries in code_map.items():
            if len(entries) > 1:
                duplicates.append({
                    "code": code,
                    "occurrences": len(entries),
                    "entries": entries
                })

        return duplicates
    finally:
        conn.close()


def reset_database():
    """Danger Zone: Clears all tables (TRUNCATE submissions & scammers)."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("TRUNCATE TABLE submissions;")
            cursor.execute("TRUNCATE TABLE scammers;")
            return {"success": True, "message": "All database tables cleared completely!"}
    finally:
        conn.close()
